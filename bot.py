import asyncio
import sqlite3
from datetime import datetime
from html import escape as html_escape
from io import BytesIO
from itertools import groupby
from pathlib import Path
from urllib.parse import urlparse

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.session.aiohttp import AiohttpSession
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.exceptions import TelegramBadRequest
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

import db
from config import BASE_URL, BOT_TOKEN, TELEGRAM_HTTP_TIMEOUT, is_admin
from link_logo import download_and_save_link_logo
from qr_image import excel_inline_qr_png, render_tracking_qr_png

router = Router()

# FSM ichida /buyruq va menyuga chiqish uchun
MAIN_MENU_TEXTS = frozenset(
    {
        "➕ Link qo'shish",
        "🎟 Promo kod qo'shish",
        "📁 Guruh qo'shish",
        "📱 QR yaratish",
        "📊 Statistika",
        "📋 Linklar",
        "🎟 Promolar",
    }
)

_SAFE_CHUNK = 3800

def main_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Link qo'shish")],
            [KeyboardButton(text="🎟 Promo kod qo'shish")],
            [
                KeyboardButton(text="📋 Linklar"),
                KeyboardButton(text="🎟 Promolar"),
                KeyboardButton(text="📁 Guruh qo'shish"),
            ],
            [KeyboardButton(text="📱 QR yaratish"), KeyboardButton(text="📊 Statistika")],
        ],
        resize_keyboard=True,
    )


class AddLinkStates(StatesGroup):
    waiting_url = State()
    optional_logo = State()


class AddPromoStates(StatesGroup):
    waiting_group = State()
    waiting_code = State()


class LogoForLinkStates(StatesGroup):
    waiting_file = State()


class EditLinkStates(StatesGroup):
    waiting_new_url = State()
    waiting_title = State()


class EditPromoStates(StatesGroup):
    waiting_code = State()


class AddGroupStates(StatesGroup):
    waiting_name = State()


def _is_valid_http_url(url: str) -> bool:
    u = urlparse(url.strip())
    return u.scheme in ("http", "https") and bool(u.netloc)


def _h(text: str) -> str:
    """Telegram HTML matn uchun xavfsiz qochirish."""
    return html_escape(text, quote=False)


def _h_attr(url: str) -> str:
    """HTML atribut (masalan <a href=\"...\">) uchun."""
    return html_escape(url.strip(), quote=True)


def _caption_tracking_link(tracking_url: str) -> str:
    """Telegramda bosiladigan havola: <code> ichidagi URL bosilmaydi."""
    u = tracking_url.strip()
    return f"<b>Havola</b>:\n<a href=\"{_h_attr(u)}\">{_h(u)}</a>"


BTN_MENU = InlineKeyboardButton(text="◀️ Menyuga", callback_data="hm:main")


def _row_menu() -> list[InlineKeyboardButton]:
    return [BTN_MENU]


def _kb_menu_only() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[_row_menu()])


def _kb_after_qr_batch(link_id: int, *, group_id: int | None = None) -> InlineKeyboardMarkup:
    if group_id is not None:
        back_cd = f"qrg:{link_id}:{group_id}"
        back_txt = "◀️ Guruhdagi promolar"
    else:
        back_cd = f"qrgl:{link_id}"
        back_txt = "◀️ Guruhlar"
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text=back_txt,
                    callback_data=back_cd,
                )
            ],
            _row_menu(),
        ]
    )


def _kb_bulk_qr_style_pick(link_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🎨 Rangli",
                    callback_data=f"qall:{link_id}:r",
                ),
                InlineKeyboardButton(
                    text="⬛ Oq-qora",
                    callback_data=f"qall:{link_id}:s",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="◀️ Guruhlar",
                    callback_data=f"qrgl:{link_id}",
                )
            ],
            _row_menu(),
        ]
    )


def _kb_bulk_qr_style_pick_group(link_id: int, group_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="🎨 Rangli",
                    callback_data=f"qallg:{link_id}:{group_id}:r",
                ),
                InlineKeyboardButton(
                    text="⬛ Oq-qora",
                    callback_data=f"qallg:{link_id}:{group_id}:s",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="◀️ Guruhdagi promolar",
                    callback_data=f"qrg:{link_id}:{group_id}",
                )
            ],
            _row_menu(),
        ]
    )


async def _send_bulk_qr_style_prompt(message: Message, link_id: int) -> None:
    await message.answer(
        "📦 <b>Barcha promo</b> — QR uslubini tanlang:",
        parse_mode=ParseMode.HTML,
        reply_markup=_kb_bulk_qr_style_pick(link_id),
    )


async def _send_bulk_qr_style_prompt_group(
    message: Message, link_id: int, group_id: int
) -> None:
    await message.answer(
        "📦 <b>Guruhdagi barcha promo</b> — QR uslubini tanlang:",
        parse_mode=ParseMode.HTML,
        reply_markup=_kb_bulk_qr_style_pick_group(link_id, group_id),
    )


def _build_tracking_links_excel(rows: list[tuple[str, str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Tracking"
    ws.append(["Guruh", "Promo kod", "Tracking havola", "QR"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row_idx, (g_name, code, url) in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=g_name)
        ws.cell(row=row_idx, column=2, value=code)
        ws.cell(row=row_idx, column=3, value=url)
        try:
            png = excel_inline_qr_png(url)
        except Exception:
            continue
        img = XLImage(BytesIO(png))
        img.width = 120
        img.height = 120
        ws.add_image(img, f"D{row_idx}")
        ws.row_dimensions[row_idx].height = 96
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 18
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _kb_link_detail(link_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✏️ O'zgartirish", callback_data=f"le:{link_id}"
                ),
                InlineKeyboardButton(
                    text="🖼 Logotipni almashtirish",
                    callback_data=f"lg:{link_id}",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="🗑 O'chirish", callback_data=f"ld:{link_id}"
                ),
            ],
            [
                InlineKeyboardButton(text="◀️ Ro'yxatga", callback_data="lb:list"),
                BTN_MENU,
            ],
        ]
    )


async def _send_link_detail(message: Message, link_id: int) -> None:
    """Logotip surati (bo'lsa) + matn + tahrir / logo / o'chirish tugmalari."""
    row = await db.get_link(link_id)
    if not row:
        await message.answer("Link topilmadi.")
        return
    u = row["url"]
    t = (row.get("title") or "").strip()
    caption = f"<b>Link #{link_id}</b>\n<b>URL</b>:\n<code>{_h(u)}</code>"
    if t:
        caption += f"\n<b>Sarlavha</b>: {_h(t)}"

    kb = _kb_link_detail(link_id)

    logo_path: Path | None = None
    lp = (row.get("logo_path") or "").strip()
    if lp:
        p = Path(lp)
        if p.is_file():
            logo_path = p
    if logo_path is None:
        disk = db.disk_path_for_link_logo(link_id)
        if disk.is_file():
            logo_path = disk

    if logo_path is not None:
        data = logo_path.read_bytes()
        photo = BufferedInputFile(data, filename="link_logo.png")
        cap = caption
        if len(cap) > 900:
            cap = cap[:897] + "…"
        await message.answer_photo(
            photo=photo,
            caption=cap,
            parse_mode=ParseMode.HTML,
            reply_markup=kb,
        )
    else:
        await message.answer(
            caption + "\n\n<i>Logotip biriktirilmagan.</i>",
            parse_mode=ParseMode.HTML,
            reply_markup=kb,
        )


async def _show_links_list_message(message: Message, intro: str) -> None:
    rows = await db.list_links()
    if not rows:
        await message.answer("Hozircha linklar yo'q.", reply_markup=_kb_menu_only())
        return
    buttons = [
        [
            InlineKeyboardButton(
                text=f"{r['id']}: {(r['url'])[:48]}…",
                callback_data=f"ll:{r['id']}",
            )
        ]
        for r in rows[:30]
    ]
    buttons.append(_row_menu())
    await message.answer(
        intro,
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


async def _link_logo_offer_links(message: Message) -> None:
    rows = await db.list_links()
    if not rows:
        await message.answer(
            "Avval «➕ Link qo'shish» bilan link qo'shing.",
            reply_markup=_kb_menu_only(),
        )
        return
    buttons = [
        [
            InlineKeyboardButton(
                text=(r.get("title") or r["url"])[:55],
                callback_data=f"lg:{r['id']}",
            )
        ]
        for r in rows[:30]
    ]
    buttons.append(_row_menu())
    await message.answer(
        "Qaysi link uchun QR markazidagi logotipni o'rnatmoqchisiz — tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


def _split_html_lines(lines: list[str]) -> list[str]:
    """Telegram 4096 cheklovidan oshmaslik uchun."""
    chunks: list[str] = []
    buf: list[str] = []
    size = 0
    for line in lines:
        add = len(line) + (1 if buf else 0)
        if buf and size + add > _SAFE_CHUNK:
            chunks.append("\n".join(buf))
            buf = [line]
            size = len(line)
        else:
            buf.append(line)
            size += add
    if buf:
        chunks.append("\n".join(buf))
    return chunks or [""]


async def _send_promo_groups_pick(
    message: Message,
    *,
    intro: str,
    prefix: str,
    include_menu: bool = True,
) -> bool:
    groups = await db.list_groups()
    if not groups:
        await message.answer(
            "Hozircha promo guruhlar yo'q. Avval bazaga kamida bitta guruh qo'shing.",
            reply_markup=_kb_menu_only() if include_menu else None,
        )
        return False
    rows = [
        [
            InlineKeyboardButton(
                text=f"📁 {g['name'][:52]}",
                callback_data=f"{prefix}:{g['id']}",
            )
        ]
        for g in groups[:60]
    ]
    if include_menu:
        rows.append(_row_menu())
    await message.answer(
        intro,
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )
    return True


def _kb_promo_detail(promo_id: int, group_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✏️ Kodni o'zgartirish",
                    callback_data=f"pec:{promo_id}",
                ),
                InlineKeyboardButton(
                    text="📁 Guruhni almashtirish",
                    callback_data=f"peg:{promo_id}",
                ),
            ],
            [
                InlineKeyboardButton(
                    text="◀️ Ro'yxatga",
                    callback_data=f"pg:{group_id}",
                ),
                BTN_MENU,
            ],
        ]
    )


async def _send_promo_detail(message: Message, promo_id: int) -> None:
    row = await db.get_promo(promo_id)
    if not row:
        await message.answer("Promo topilmadi.", reply_markup=_kb_menu_only())
        return
    text = (
        f"<b>Promo #{promo_id}</b>\n"
        f"<b>Kod</b>: <code>{_h(row['code'])}</code>\n"
        f"<b>Guruh</b>: {_h(row['group_name'])}"
    )
    await message.answer(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=_kb_promo_detail(promo_id, int(row["group_id"])),
    )


async def _send_promos_in_group(message: Message, group_id: int) -> None:
    group = await db.get_group(group_id)
    if not group:
        await message.answer("Guruh topilmadi.", reply_markup=_kb_menu_only())
        return
    promos = await db.list_promos_by_group(group_id)
    if not promos:
        await message.answer(
            f"📁 <b>{_h(group['name'])}</b>\n\nBu guruhda promolar yo'q.",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="◀️ Guruhlarga", callback_data="pg:list")],
                    _row_menu(),
                ]
            ),
        )
        return
    buttons = [
        [
            InlineKeyboardButton(
                text=p["code"][:60],
                callback_data=f"pp:{group_id}:{p['id']}",
            )
        ]
        for p in promos[:60]
    ]
    buttons.append(
        [
            InlineKeyboardButton(text="◀️ Guruhlarga", callback_data="pg:list"),
            BTN_MENU,
        ]
    )
    await message.answer(
        f"📁 <b>{_h(group['name'])}</b>\n\nKerakli promoni tanlang:",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("Bu bot faqat administratorlar uchun.")
        return
    await message.answer("Salom!", reply_markup=main_kb())


@router.message(Command("help"))
async def cmd_help(message: Message) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await message.answer(
        "1) Avval <b>Link</b> va <b>Promo guruh/kod</b> qo'shing.\n"
        "2) <b>📁 Guruh qo'shish</b> — yangi promo guruh nomi (takroriysiz).\n"
        "3) <b>QR yaratish</b> — link, keyin guruh, promo yoki guruhdagi/barcha promo "
        "uchun QR; tracking havolalar Excel ham mavjud.\n"
        "4) <b>Statistika</b> — guruh va yuklanishlar, Excel export.\n"
        "5) Har bir link uchun QR markazidagi logotip — link qo'shganda yoki "
        "<b>📋 Linklar</b> → link kartochkasidagi <b>Logotipni almashtirish</b>.\n"
        "6) <b>📋 Linklar</b> — tahrir, logotip va o'chirish ham shu yerda.\n\n"
        f"Tracking URL ko'rinishi: <code>{_h(BASE_URL)}/r/token</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=_kb_menu_only(),
    )


@router.message(Command("cancel"))
async def cancel(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await state.clear()
    await message.answer("Bekor qilindi.", reply_markup=main_kb())


def _deny(message: Message) -> bool:
    if message.from_user and is_admin(message.from_user.id):
        return False
    return True


@router.message(F.text == "➕ Link qo'shish")
async def add_link_prompt(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await state.set_state(AddLinkStates.waiting_url)
    await message.answer(
        "To'liq URL yuboring (https://...).\nBekor qilish: /cancel",
        reply_markup=ReplyKeyboardRemove(),
    )


@router.message(F.text == "📁 Guruh qo'shish")
async def add_group_prompt(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await state.set_state(AddGroupStates.waiting_name)
    await message.answer(
        "Yangi promo <b>guruh nomini</b> yuboring.\nBekor qilish: /cancel",
        parse_mode=ParseMode.HTML,
        reply_markup=ReplyKeyboardRemove(),
    )


@router.message(AddGroupStates.waiting_name, F.text & ~F.text.startswith("/"))
async def add_group_save(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    name = (message.text or "").strip()
    if len(name) < 2:
        await message.answer("Nom juda qisqa.")
        return
    try:
        gid = await db.add_group(name)
    except sqlite3.IntegrityError:
        await message.answer("Bu nom bilan guruh allaqachon mavjud.")
        return
    await state.clear()
    await message.answer(
        f"✅ Guruh saqlandi (id: <code>{gid}</code>).",
        parse_mode=ParseMode.HTML,
        reply_markup=main_kb(),
    )


@router.message(AddLinkStates.waiting_url, F.text.in_(MAIN_MENU_TEXTS))
@router.message(AddLinkStates.optional_logo, F.text.in_(MAIN_MENU_TEXTS))
@router.message(AddPromoStates.waiting_group, F.text.in_(MAIN_MENU_TEXTS))
@router.message(AddPromoStates.waiting_code, F.text.in_(MAIN_MENU_TEXTS))
@router.message(AddGroupStates.waiting_name, F.text.in_(MAIN_MENU_TEXTS))
@router.message(LogoForLinkStates.waiting_file, F.text.in_(MAIN_MENU_TEXTS))
@router.message(EditLinkStates.waiting_new_url, F.text.in_(MAIN_MENU_TEXTS))
@router.message(EditLinkStates.waiting_title, F.text.in_(MAIN_MENU_TEXTS))
@router.message(EditPromoStates.waiting_code, F.text.in_(MAIN_MENU_TEXTS))
async def fsm_to_main_menu(message: Message, state: FSMContext) -> None:
    """FSM ichida pastki menyu tugmalari — holatni tozalaydi va buyruqni bajaradi."""
    if _deny(message):
        return
    await state.clear()
    text = message.text or ""
    if text == "📊 Statistika":
        await stats_cmd(message)
    elif text == "📋 Linklar":
        await list_links_cmd(message)
    elif text == "🎟 Promolar":
        await list_promos_cmd(message)
    elif text == "📱 QR yaratish":
        await qr_pick_link(message)
    elif text == "➕ Link qo'shish":
        await add_link_prompt(message, state)
    elif text == "🎟 Promo kod qo'shish":
        await add_promo_prompt(message, state)
    elif text == "📁 Guruh qo'shish":
        await add_group_prompt(message, state)


@router.callback_query(F.data == "hm:main")
async def callback_home_menu(callback: CallbackQuery) -> None:
    """Inline menyudan pastki klaviaturaga qaytish."""
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await callback.message.answer("Pastki menyu:", reply_markup=main_kb())
    await callback.answer()


@router.callback_query(F.data == "lb:list")
async def link_list_back(callback: CallbackQuery) -> None:
    """Link kartochkasidan ro'yxatga qaytish."""
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _show_links_list_message(
            callback.message,
            "📋 <b>Linklar</b>\n\nKerakli linkni tanlang:",
        )
    await callback.answer()


@router.callback_query(F.data.startswith("ll:"))
async def link_detail_open(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_link_detail(callback.message, link_id)
    await callback.answer()


@router.callback_query(F.data.startswith("le:"))
async def link_edit_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    await state.set_state(EditLinkStates.waiting_new_url)
    await state.update_data(edit_link_id=link_id)
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await callback.message.answer(
            "Yangi <b>to'liq URL</b> yuboring (<code>https://...</code>).\nBekor: /cancel",
            parse_mode=ParseMode.HTML,
        )
    await callback.answer()


@router.message(EditLinkStates.waiting_new_url, F.text & ~F.text.startswith("/"))
async def link_edit_save_url(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    data = await state.get_data()
    lid = data.get("edit_link_id")
    if not isinstance(lid, int):
        await state.clear()
        return
    url = (message.text or "").strip()
    if not _is_valid_http_url(url):
        await message.answer("Noto'g'ri URL. https:// bilan qayta yuboring.")
        return
    await db.update_link_fields(lid, url=url)
    await state.set_state(EditLinkStates.waiting_title)
    await message.answer(
        "URL yangilandi.\n\n"
        "<b>Sarlavha</b> (ixtiyoriy, ro'yxatda qisqa ko'rinish uchun) yuboring yoki /skip",
        parse_mode=ParseMode.HTML,
    )


@router.message(EditLinkStates.waiting_title, Command("skip"))
async def link_edit_skip_title(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await state.clear()
    await message.answer("Tahrir yakunlandi.", reply_markup=main_kb())


@router.message(EditLinkStates.waiting_title, F.text)
async def link_edit_save_title(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    data = await state.get_data()
    lid = data.get("edit_link_id")
    if not isinstance(lid, int):
        await state.clear()
        return
    title = (message.text or "").strip()
    await db.update_link_fields(lid, title=title if title else None)
    await state.clear()
    await message.answer("Sarlavha saqlandi.", reply_markup=main_kb())


@router.callback_query(F.data.startswith("ld:"))
async def link_delete_ask(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✅ Ha, o'chirish",
                    callback_data=f"ldc:{link_id}",
                )
            ],
            [InlineKeyboardButton(text="❌ Yo'q", callback_data=f"ll:{link_id}")],
            _row_menu(),
        ]
    )
    if callback.message:
        await callback.message.answer(
            f"<b>Link #{link_id}</b> va unga bog'langan barcha tracking/QR yozuvlari "
            f"o'chiriladi. Davom etasizmi?",
            parse_mode=ParseMode.HTML,
            reply_markup=kb,
        )
    await callback.answer()


@router.callback_query(F.data.startswith("ldc:"))
async def link_delete_do(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    ok = await db.delete_link(link_id)
    if not ok:
        await callback.answer("Link topilmadi", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await callback.message.answer(
            f"✅ Link <code>{link_id}</code> o'chirildi.",
            parse_mode=ParseMode.HTML,
        )
    await callback.answer("O'chirildi")


@router.message(AddLinkStates.waiting_url, F.text & ~F.text.startswith("/"))
async def add_link_save(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    url = (message.text or "").strip()
    if not _is_valid_http_url(url):
        await message.answer("Noto'g'ri URL. https:// bilan qayta yuboring.")
        return
    lid = await db.add_link(url)
    await state.update_data(pending_link_id=lid)
    await state.set_state(AddLinkStates.optional_logo)
    await message.answer(
        f"✅ Link saqlandi (id: <code>{lid}</code>).\n\n"
        "QR markazidagi <b>logotip</b> uchun surat yoki PNG/JPEG <b>dokument</b> yuboring.\n"
        "O'tkazib yuborish: /skip",
        parse_mode=ParseMode.HTML,
        reply_markup=ReplyKeyboardRemove(),
    )


@router.message(AddLinkStates.optional_logo, Command("skip"))
async def add_link_skip_logo(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await state.clear()
    await message.answer("Logotipsiz saqlandi.", reply_markup=main_kb())


@router.message(AddLinkStates.optional_logo, F.text)
async def add_link_logo_hint(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await message.answer(
        "Logotipni <b>surat</b> yoki <b>dokument</b> (PNG/JPEG) sifatida yuboring, yoki /skip",
        parse_mode=ParseMode.HTML,
    )


@router.message(AddLinkStates.optional_logo, F.document | F.photo)
async def add_link_save_logo(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    data = await state.get_data()
    lid = data.get("pending_link_id")
    if not isinstance(lid, int):
        await state.clear()
        return
    try:
        await download_and_save_link_logo(message, lid)
    except ValueError as e:
        await message.answer(_h(str(e)))
        return
    await state.clear()
    await _send_link_detail(message, lid)
    await message.answer("✅", reply_markup=main_kb())


@router.message(F.text == "🎟 Promo kod qo'shish")
async def add_promo_prompt(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    await state.set_state(AddPromoStates.waiting_group)
    await message.answer(
        "Yangi promo uchun avval guruhni tanlang:",
        reply_markup=ReplyKeyboardRemove(),
    )
    await _send_promo_groups_pick(
        message,
        intro="📁 <b>Promo guruhlar</b>\n\nTanlang:",
        prefix="apg",
        include_menu=False,
    )


@router.callback_query(F.data.startswith("apg:"))
async def add_promo_pick_group(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        group_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    group = await db.get_group(group_id)
    if not group:
        await callback.answer("Guruh topilmadi", show_alert=True)
        return
    await state.update_data(add_promo_group_id=group_id)
    await state.set_state(AddPromoStates.waiting_code)
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await callback.message.answer(
            f"📁 Tanlandi: <b>{_h(group['name'])}</b>\n\n"
            "Promo kod matnini yuboring (masalan: SUMMER2026).",
            parse_mode=ParseMode.HTML,
        )
    await callback.answer()


@router.message(AddPromoStates.waiting_code, F.text & ~F.text.startswith("/"))
async def add_promo_save(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    data = await state.get_data()
    group_id = data.get("add_promo_group_id")
    if not isinstance(group_id, int):
        await state.clear()
        await message.answer("Guruh tanlanmadi. Qaytadan boshlang.", reply_markup=main_kb())
        return
    code = (message.text or "").strip()
    if len(code) < 2:
        await message.answer("Kod juda qisqa.")
        return
    try:
        pid = await db.add_promo(code, group_id)
    except ValueError:
        await state.clear()
        await message.answer("Guruh topilmadi. Qaytadan urinib ko'ring.", reply_markup=main_kb())
        return
    except sqlite3.IntegrityError:
        await message.answer("Bu promo kod allaqachon mavjud.")
        return
    await state.clear()
    await message.answer(f"✅ Promo saqlandi (id: {pid}).", reply_markup=main_kb())


@router.message(F.text == "📋 Linklar")
async def list_links_cmd(message: Message) -> None:
    if _deny(message):
        return
    await _show_links_list_message(
        message,
        "📋 <b>Linklar</b>\n\nTanlang — keyin logotip va tahrir tugmalari chiqadi:",
    )


@router.message(F.text == "🎟 Promolar")
async def list_promos_cmd(message: Message) -> None:
    if _deny(message):
        return
    await _send_promo_groups_pick(
        message,
        intro="🎟 <b>Promolar</b>\n\nAvval guruhni tanlang:",
        prefix="pg",
    )


@router.callback_query(F.data == "pg:list")
async def promo_groups_list_back(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_promo_groups_pick(
            callback.message,
            intro="🎟 <b>Promolar</b>\n\nAvval guruhni tanlang:",
            prefix="pg",
        )
    await callback.answer()


@router.callback_query(F.data.startswith("pg:"))
async def promo_group_open(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        group_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_promos_in_group(callback.message, group_id)
    await callback.answer()


@router.callback_query(F.data.startswith("pp:"))
async def promo_detail_open(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Xato", show_alert=True)
        return
    try:
        promo_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_promo_detail(callback.message, promo_id)
    await callback.answer()


@router.callback_query(F.data.startswith("pec:"))
async def promo_edit_code_start(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        promo_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    promo = await db.get_promo(promo_id)
    if not promo:
        await callback.answer("Promo topilmadi", show_alert=True)
        return
    await state.set_state(EditPromoStates.waiting_code)
    await state.update_data(edit_promo_id=promo_id)
    if callback.message:
        await callback.message.answer(
            "Yangi promo kodni yuboring.\nBekor: /cancel",
        )
    await callback.answer()


@router.message(EditPromoStates.waiting_code, F.text & ~F.text.startswith("/"))
async def promo_edit_code_save(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    data = await state.get_data()
    promo_id = data.get("edit_promo_id")
    if not isinstance(promo_id, int):
        await state.clear()
        return
    code = (message.text or "").strip()
    if len(code) < 2:
        await message.answer("Kod juda qisqa.")
        return
    try:
        ok = await db.update_promo_fields(promo_id, code=code)
    except sqlite3.IntegrityError:
        await message.answer("Bu promo kod allaqachon mavjud.")
        return
    await state.clear()
    if not ok:
        await message.answer("Promo topilmadi.", reply_markup=main_kb())
        return
    await _send_promo_detail(message, promo_id)
    await message.answer("✅", reply_markup=main_kb())


@router.callback_query(F.data.startswith("peg:"))
async def promo_edit_group_start(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        promo_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    promo = await db.get_promo(promo_id)
    if not promo:
        await callback.answer("Promo topilmadi", show_alert=True)
        return
    groups = await db.list_groups()
    buttons = [
        [
            InlineKeyboardButton(
                text=f"📁 {g['name'][:52]}",
                callback_data=f"pegs:{promo_id}:{g['id']}",
            )
        ]
        for g in groups[:60]
    ]
    buttons.append(
        [
            InlineKeyboardButton(
                text="◀️ Orqaga",
                callback_data=f"pp:{promo['group_id']}:{promo_id}",
            ),
            BTN_MENU,
        ]
    )
    if callback.message:
        await callback.message.answer(
            "Promoning yangi guruhini tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
        )
    await callback.answer()


@router.callback_query(F.data.startswith("pegs:"))
async def promo_edit_group_save(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Xato", show_alert=True)
        return
    try:
        promo_id = int(parts[1])
        group_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    ok = await db.update_promo_fields(promo_id, group_id=group_id)
    if not ok:
        await callback.answer("Promo yoki guruh topilmadi", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_promo_detail(callback.message, promo_id)
    await callback.answer("Saqlandi")


def _kb_stats_root() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="📁 Guruh bo'yicha ko'rish",
                    callback_data="stat:glist",
                )
            ],
            [
                InlineKeyboardButton(
                    text="📥 Excel yuklash",
                    callback_data="stat:xlsx",
                )
            ],
            _row_menu(),
        ]
    )


def _kb_stats_group_detail() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="◀️ Guruhlar ro'yxati",
                    callback_data="stat:glist",
                )
            ],
            _row_menu(),
        ]
    )


async def _show_stats_root(message: Message) -> None:
    groups = await db.list_groups()
    if not groups:
        await message.answer(
            "Hozircha promo guruhlar yo'q.",
            reply_markup=_kb_menu_only(),
        )
        return
    promos = await db.list_promos()
    if not promos:
        await message.answer(
            "Statistika uchun avval <b>promo kod</b> qo'shing.",
            parse_mode=ParseMode.HTML,
            reply_markup=_kb_menu_only(),
        )
        return
    totals = await db.stats_group_totals_desc()
    lines = []
    total_clicks = 0
    for row in totals:
        clicks = int(row.get("clicks") or 0)
        total_clicks += clicks
        raw_name = str(row.get("group_name") or "")
        lines.append(f"{raw_name} — {clicks:,}".replace(",", " "))
    stats_block = "\n\n".join(lines) if lines else "Hozircha ma'lumot yo'q."
    await message.answer(
        "📊 <b>Yuklanishlar statistikasi</b>\n\n"
        f"{_h(stats_block)}\n\n"
        f"<b>Jami:</b> {_h(f'{total_clicks:,}'.replace(',', ' '))} ta yuklanish",
        parse_mode=ParseMode.HTML,
        reply_markup=_kb_stats_root(),
    )


async def _show_stats_group_list(message: Message) -> None:
    await _send_promo_groups_pick(
        message,
        intro="📊 <b>Statistika</b>\n\nGuruhni tanlang:",
        prefix="statg",
    )

async def _send_stats_for_group_detail(message: Message, group_id: int) -> None:
    group = await db.get_group(group_id)
    if not group:
        await message.answer("Guruh topilmadi.", reply_markup=_kb_menu_only())
        return
    stats = await db.stats_for_group(group_id)
    head = (
        "📊 <b>Statistika</b>\n"
        f"<b>Guruh</b>: {_h(group['name'])}\n\n"
        "<b>Promo va jami yuklanishlar</b>\n"
    )
    if not stats:
        promo_lines = ["<i>Hozircha promo kodlar yo'q.</i>"]
        total_clicks = 0
    else:
        total_clicks = sum(int(s["clicks"] or 0) for s in stats)
        promo_lines = [
            f"• <code>{_h(s['promo_code'])}</code> — "
            f"<b>{int(s['clicks'] or 0)}</b> yuklanish"
            for s in stats
        ]
        promo_lines.append("")
        promo_lines.append(f"<b>Jami</b>: <b>{total_clicks}</b> yuklanish")
    parts = _split_html_lines(promo_lines)
    kb = _kb_stats_group_detail()
    messages_out: list[str] = []
    for i, p in enumerate(parts):
        prefix = head if i == 0 else "📊 <i>Davomi</i>\n\n"
        messages_out.append(prefix + p)
    for i, text in enumerate(messages_out):
        await message.answer(
            text,
            parse_mode=ParseMode.HTML,
            reply_markup=kb if i == len(messages_out) - 1 else None,
        )


def _build_stats_excel(
    promo_rows: list[dict[str, object]],
    group_rows: list[dict[str, object]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Statistika"
    ws.append(["Guruh", "Promo kod", "Yuklanishlar"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    grand_total = 0
    for group_name, items in groupby(
        promo_rows, key=lambda r: str(r.get("group_name") or "")
    ):
        group_sum = 0
        for item in items:
            clicks = int(item.get("clicks") or 0)
            ws.append(
                [
                    group_name,
                    str(item.get("promo_code") or ""),
                    clicks,
                ]
            )
            group_sum += clicks
        ws.append([f"{group_name} jami", "", group_sum])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
        ws.append([])
        grand_total += group_sum

    ws.append(["Jami", "", grand_total])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14

    ws2 = wb.create_sheet("Guruhlar")
    ws2.append(["Guruh", "Prioritet", "Promo soni", "Yuklanishlar"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    total_promos = 0
    total_clicks = 0
    for row in group_rows:
        promo_count = int(row.get("promo_count") or 0)
        clicks = int(row.get("clicks") or 0)
        total_promos += promo_count
        total_clicks += clicks
        ws2.append(
            [
                str(row.get("group_name") or ""),
                int(row.get("priority") or 0),
                promo_count,
                clicks,
            ]
        )
    ws2.append(["Jami", "", total_promos, total_clicks])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=3).font = Font(bold=True)
    ws2.cell(row=ws2.max_row, column=4).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@router.callback_query(F.data == "stat:glist")
async def stats_back_to_groups(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _show_stats_group_list(callback.message)
    await callback.answer()


@router.callback_query(F.data.startswith("statg:"))
async def stats_open_group(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        group_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_group(group_id):
        await callback.answer("Guruh topilmadi", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_stats_for_group_detail(callback.message, group_id)
    await callback.answer()


@router.callback_query(F.data == "stat:xlsx")
async def stats_export_excel(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    rows = await db.stats_summary_by_group()
    group_rows = await db.stats_group_totals_desc()
    if not rows:
        await callback.answer("Statistika uchun ma'lumot yo'q", show_alert=True)
        return
    content = _build_stats_excel(rows, group_rows)
    filename = f"statistika_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    file = BufferedInputFile(content, filename=filename)
    if callback.message:
        await callback.message.answer_document(
            file,
            caption="📥 Guruhlar kesimidagi statistika (Excel).",
        )
    await callback.answer()


@router.message(F.text == "📊 Statistika")
async def stats_cmd(message: Message) -> None:
    if _deny(message):
        return
    await _show_stats_root(message)


async def _send_qr_link_pick(target: Message) -> bool:
    """QR oqimi: link tanlash. False bo'lsa shartlar bajarilmagan."""
    links = await db.list_links()
    if not links:
        await target.answer(
            "Avval «Link qo'shish» orqali link qo'shing.",
            reply_markup=_kb_menu_only(),
        )
        return False
    groups = await db.list_groups()
    if not groups:
        await target.answer(
            "Avval promo <b>guruh</b> yarating yoki «📁 Guruh qo'shish»dan qo'shing.",
            parse_mode=ParseMode.HTML,
            reply_markup=_kb_menu_only(),
        )
        return False
    promos = await db.list_promos()
    if not promos:
        await target.answer(
            "Avval «Promo kod qo'shish» orqali promo qo'shing.",
            reply_markup=_kb_menu_only(),
        )
        return False
    buttons = [
        [
            InlineKeyboardButton(
                text=(l.get("title") or l["url"])[:60],
                callback_data=f"ql:{l['id']}",
            )
        ]
        for l in links[:30]
    ]
    buttons.append(_row_menu())
    await target.answer(
        "Tracking uchun linkni tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    return True


async def _send_qr_group_pick(target: Message, link_id: int) -> None:
    groups = await db.list_groups()
    if not groups:
        await target.answer(
            "Promo guruhlar yo'q.",
            reply_markup=_kb_menu_only(),
        )
        return
    rows: list[list[InlineKeyboardButton]] = []
    for g in groups[:55]:
        rows.append(
            [
                InlineKeyboardButton(
                    text=f"📁 {g['name'][:52]}",
                    callback_data=f"qrg:{link_id}:{g['id']}",
                )
            ]
        )
    rows.append(
        [
            InlineKeyboardButton(
                text="📦 Barcha promo uchun QR",
                callback_data=f"qbulk:{link_id}",
            ),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton(
                text="📥 Excel — barcha promo (havolalar)",
                callback_data=f"qrxlall:{link_id}",
            ),
        ]
    )
    rows.append(
        [
            InlineKeyboardButton(text="◀️ Link tanlash", callback_data="qr:bl"),
            BTN_MENU,
        ]
    )
    await target.answer(
        "Avval <b>promo guruhini</b> tanlang "
        "(yoki barcha promo uchun QR / Excel):",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=rows),
    )


async def _send_qr_promo_pick_for_group(
    target: Message, link_id: int, group_id: int
) -> None:
    group = await db.get_group(group_id)
    if not group:
        await target.answer("Guruh topilmadi.", reply_markup=_kb_menu_only())
        return
    promos = await db.list_promos_by_group(group_id)
    buttons: list[list[InlineKeyboardButton]] = [
        [
            InlineKeyboardButton(
                text="📦 Shu guruhdagi barcha promo uchun QR",
                callback_data=f"qbulkg:{link_id}:{group_id}",
            ),
        ],
        [
            InlineKeyboardButton(
                text="📥 Excel — shu guruh (havolalar)",
                callback_data=f"qrxlg:{link_id}:{group_id}",
            ),
        ],
    ]
    for p in promos[:50]:
        buttons.append(
            [
                InlineKeyboardButton(
                    text=p["code"][:58],
                    callback_data=f"qp:{link_id}:{p['id']}:{group_id}",
                )
            ]
        )
    buttons.append(
        [
            InlineKeyboardButton(
                text="◀️ Guruhlar",
                callback_data=f"qrgl:{link_id}",
            ),
            BTN_MENU,
        ]
    )
    await target.answer(
        f"📁 <b>{_h(group['name'])}</b>\n\n"
        "Promo tanlang, guruhdagi barcha uchun QR yoki Excel:",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


async def _answer_tracking_excel(
    message: Message,
    link_id: int,
    promos: list[dict[str, object]],
    *,
    caption: str,
) -> None:
    if not promos:
        await message.answer("Promo kodlar yo'q.")
        return
    link_row = await db.get_link(link_id)
    if not link_row:
        await message.answer("Link topilmadi.")
        return
    ids = [int(p["id"]) for p in promos]
    token_map = await db.ensure_track_tokens_for_promos(link_id, ids)
    rows_data: list[tuple[str, str, str]] = []
    for p in promos:
        pid = int(p["id"])
        token = token_map[pid]
        url = f"{BASE_URL}/r/{token}"
        rows_data.append(
            (
                str(p.get("group_name") or ""),
                str(p.get("code") or ""),
                url,
            )
        )
    content = await asyncio.to_thread(_build_tracking_links_excel, rows_data)
    filename = f"qr_tracking_link{link_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await message.answer_document(
        BufferedInputFile(content, filename=filename),
        caption=caption,
    )


async def _deliver_bulk_promo_qrs(
    message: Message,
    link_id: int,
    style: str,
    *,
    group_id: int | None = None,
) -> None:
    if group_id is not None:
        promos = await db.list_promos_by_group(group_id)
    else:
        promos = await db.list_promos()
    if not promos:
        await message.answer(
            "Promo kodlar yo'q.",
            reply_markup=_kb_menu_only(),
        )
        return
    link_row = await db.get_link(link_id)
    if not link_row:
        await message.answer("Link topilmadi.", reply_markup=_kb_menu_only())
        return
    logo_override: str | None = None
    if link_row:
        lp = (link_row.get("logo_path") or "").strip()
        if lp and Path(lp).is_file():
            logo_override = lp
    uslub = "Oddiy (oq-qora)" if style == "simple" else "Rangli"
    total = len(promos)
    await message.answer(
        f"📦 <b>{total}</b> ta promo uchun QR yuborilmoqda "
        f"(<i>{uslub}</i>)…",
        parse_mode=ParseMode.HTML,
    )
    for i, p in enumerate(promos):
        promo_id = p["id"]
        code = p["code"]
        token = await db.create_track_entry(link_id, promo_id)
        tracking_url = f"{BASE_URL}/r/{token}"
        png_bytes = render_tracking_qr_png(
            tracking_url, logo_path=logo_override, style=style
        )
        safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in code)[
            :40
        ]
        file = BufferedInputFile(png_bytes, filename=f"qr_{safe_name or promo_id}.png")
        last = i == total - 1
        await message.answer_photo(
            photo=file,
            caption=(
                f"<b>Promo</b>: <code>{_h(code)}</code>\n"
                f"<b>Uslub</b>: {uslub}\n"
                f"{_caption_tracking_link(tracking_url)}"
            ),
            parse_mode=ParseMode.HTML,
            reply_markup=_kb_after_qr_batch(link_id, group_id=group_id)
            if last
            else None,
        )
        if not last:
            await asyncio.sleep(1)


@router.message(F.text == "📱 QR yaratish")
async def qr_pick_link(message: Message) -> None:
    if _deny(message):
        return
    await _send_qr_link_pick(message)


@router.callback_query(F.data == "qr:bl")
async def qr_back_to_links(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_qr_link_pick(callback.message)
    await callback.answer()


@router.callback_query(F.data.startswith("qrgl:"))
async def qr_back_to_group_pick(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_qr_group_pick(callback.message, link_id)
    await callback.answer()


@router.callback_query(F.data.startswith("qrg:"))
async def qr_open_group_for_promos(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
        group_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id) or not await db.get_group(group_id):
        await callback.answer("Ma'lumot topilmadi", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_qr_promo_pick_for_group(callback.message, link_id, group_id)
    await callback.answer()


@router.callback_query(F.data.startswith("qbulkg:"))
async def qr_bulk_group_open_style_menu(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
        group_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id) or not await db.get_group(group_id):
        await callback.answer("Ma'lumot topilmadi", show_alert=True)
        return
    promos = await db.list_promos_by_group(group_id)
    if not promos:
        await callback.answer("Guruhda promo yo'q", show_alert=True)
        return
    await callback.answer()
    if callback.message:
        await _send_bulk_qr_style_prompt_group(
            callback.message, link_id, group_id
        )


@router.callback_query(F.data.startswith("qrxlall:"))
async def qr_excel_all_promos(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    promos = await db.list_promos()
    if not promos:
        await callback.answer("Promo yo'q", show_alert=True)
        return
    await callback.answer("Excel tayyorlanmoqda…")
    if callback.message:
        await _answer_tracking_excel(
            callback.message,
            link_id,
            promos,
            caption="📥 Barcha promo uchun tracking havolalar (Excel).",
        )


@router.callback_query(F.data.startswith("qrxlg:"))
async def qr_excel_one_group(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 3:
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
        group_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id) or not await db.get_group(group_id):
        await callback.answer("Ma'lumot topilmadi", show_alert=True)
        return
    promos = await db.list_promos_by_group(group_id)
    if not promos:
        await callback.answer("Guruhda promo yo'q", show_alert=True)
        return
    await callback.answer("Excel tayyorlanmoqda…")
    if callback.message:
        g = await db.get_group(group_id)
        name = (g or {}).get("name") or ""
        await _answer_tracking_excel(
            callback.message,
            link_id,
            promos,
            caption=f"📥 Guruhdagi promo havolalar — {name}",
        )


@router.callback_query(F.data.startswith("qbulk:"))
async def qr_bulk_open_style_menu(callback: CallbackQuery) -> None:
    """Barcha promo QR — avval rangli / oq-qora tanlash."""
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    await callback.answer()
    if callback.message:
        await _send_bulk_qr_style_prompt(callback.message, link_id)


@router.callback_query(F.data.startswith("qall:"))
async def qr_all_promos_bulk(callback: CallbackQuery) -> None:
    """Tanlangan link uchun bazadagi barcha promolar bo'yicha QR yuborish."""
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) < 2 or parts[0] != "qall":
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    if len(parts) < 3 or parts[2] not in ("s", "r"):
        await callback.answer()
        if callback.message:
            await _send_bulk_qr_style_prompt(callback.message, link_id)
        return
    mode = parts[2]
    style = "simple" if mode == "s" else "styled"
    await callback.answer("QRlar yuborilmoqda…")
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _deliver_bulk_promo_qrs(callback.message, link_id, style)


@router.callback_query(F.data.startswith("qallg:"))
async def qr_all_promos_in_group_bulk(callback: CallbackQuery) -> None:
    """Tanlangan link + guruh uchun barcha promo QR."""
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 4 or parts[0] != "qallg":
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
        group_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    mode = parts[3]
    if not await db.get_link(link_id) or not await db.get_group(group_id):
        await callback.answer("Ma'lumot topilmadi", show_alert=True)
        return
    if mode not in ("s", "r"):
        await callback.answer()
        if callback.message:
            await _send_bulk_qr_style_prompt_group(
                callback.message, link_id, group_id
            )
        return
    style = "simple" if mode == "s" else "styled"
    promos = await db.list_promos_by_group(group_id)
    if not promos:
        await callback.answer("Guruhda promo yo'q", show_alert=True)
        return
    await callback.answer("QRlar yuborilmoqda…")
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _deliver_bulk_promo_qrs(
            callback.message, link_id, style, group_id=group_id
        )


@router.callback_query(F.data.startswith("lg:"))
async def link_logo_callback(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if not await db.get_link(link_id):
        await callback.answer("Link topilmadi", show_alert=True)
        return
    await state.set_state(LogoForLinkStates.waiting_file)
    await state.update_data(logo_link_id=link_id)
    if callback.message:
        await callback.message.answer(
            f"Link id <code>{link_id}</code>: logotip yuboring (surat yoki PNG dokument).\n"
            "Bekor: /cancel",
            parse_mode=ParseMode.HTML,
        )
    await callback.answer()


@router.message(LogoForLinkStates.waiting_file, F.document | F.photo)
async def link_logo_save_for_existing(message: Message, state: FSMContext) -> None:
    if _deny(message):
        return
    data = await state.get_data()
    lid = data.get("logo_link_id")
    if not isinstance(lid, int):
        await state.clear()
        return
    try:
        await download_and_save_link_logo(message, lid)
    except ValueError as e:
        await message.answer(_h(str(e)))
        return
    await state.clear()
    await _send_link_detail(message, lid)
    await message.answer("✅", reply_markup=main_kb())


@router.callback_query(F.data.startswith("ql:"))
async def qr_pick_promo(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    try:
        link_id = int(callback.data.split(":")[1])
    except (IndexError, ValueError):
        await callback.answer("Xato", show_alert=True)
        return
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await _send_qr_group_pick(callback.message, link_id)
    await callback.answer()


@router.callback_query(F.data.startswith("qp:"))
async def qr_choose_style(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) not in (3, 4) or parts[0] != "qp":
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
        promo_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    if len(parts) == 4:
        try:
            group_id = int(parts[3])
        except ValueError:
            await callback.answer("Xato", show_alert=True)
            return
        back_cd = f"qrg:{link_id}:{group_id}"
        back_txt = "◀️ Guruhdagi promolar"
    else:
        back_cd = f"qrgl:{link_id}"
        back_txt = "◀️ Guruhlar"

    style_row = [
        InlineKeyboardButton(
            text="⬛ Oddiy (oq-qora)",
            callback_data=f"qst:{link_id}:{promo_id}:s",
        ),
        InlineKeyboardButton(
            text="🎨 Rangli",
            callback_data=f"qst:{link_id}:{promo_id}:r",
        ),
    ]
    if callback.message:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except TelegramBadRequest:
            pass
        await callback.message.answer(
            "QR kod turini tanlang:",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    style_row,
                    [
                        InlineKeyboardButton(
                            text=back_txt,
                            callback_data=back_cd,
                        ),
                    ],
                    _row_menu(),
                ]
            ),
        )
    await callback.answer()


@router.callback_query(F.data.startswith("qst:"))
async def qr_build(callback: CallbackQuery) -> None:
    if not callback.from_user or not is_admin(callback.from_user.id):
        await callback.answer("Ruxsat yo'q", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 4:
        await callback.answer("Xato", show_alert=True)
        return
    try:
        link_id = int(parts[1])
        promo_id = int(parts[2])
    except ValueError:
        await callback.answer("Xato", show_alert=True)
        return
    mode = parts[3]
    if mode not in ("s", "r"):
        await callback.answer("Xato", show_alert=True)
        return
    style = "simple" if mode == "s" else "styled"

    token = await db.create_track_entry(link_id, promo_id)
    tracking_url = f"{BASE_URL}/r/{token}"
    promo_label = await db.get_promo_code_by_id(promo_id) or "?"

    link_row = await db.get_link(link_id)
    logo_override: str | None = None
    if link_row:
        lp = (link_row.get("logo_path") or "").strip()
        if lp and Path(lp).is_file():
            logo_override = lp

    png_bytes = render_tracking_qr_png(
        tracking_url, logo_path=logo_override, style=style
    )
    file = BufferedInputFile(png_bytes, filename="promo_qr.png")

    uslub = "Oddiy (oq-qora)" if style == "simple" else "Rangli"

    if not callback.message:
        await callback.answer("Xabar topilmadi", show_alert=True)
        return
    await callback.message.answer_photo(
        photo=file,
        caption=(
            f"<b>Promo kod</b>: <code>{_h(promo_label)}</code>\n"
            f"<b>Uslub</b>: {uslub}\n"
            f"{_caption_tracking_link(tracking_url)}"
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=_kb_menu_only(),
    )
    try:
        await callback.message.edit_reply_markup(reply_markup=None)
    except TelegramBadRequest:
        pass
    await callback.answer("Tayyor")


def get_dispatcher() -> Dispatcher:
    dp = Dispatcher()
    dp.include_router(router)
    return dp


def make_bot() -> Bot:
    """Telegram API uchun uzoq timeout — tarmoq uzilishi / Windows ClientOSError kamayishi mumkin."""
    session = AiohttpSession(timeout=TELEGRAM_HTTP_TIMEOUT)
    return Bot(BOT_TOKEN, session=session)


async def run_bot_polling() -> None:
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN .env faylida yo'q")
    bot = make_bot()
    dp = get_dispatcher()
    await dp.start_polling(bot)
